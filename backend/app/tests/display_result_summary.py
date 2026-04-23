#!/usr/bin/env python3
import sys
sys.path.insert(0, 'C:\\Users\\ryanj\\Mark-UP\\backend')

from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\ryanj\\Mark-UP\\backend\\data\\outputs\\result.xlsx')
ws = wb.active

print("=== FINAL RESULT.XLSX SUMMARY ===\n")
print("Student Data with Totals:\n")

for i in range(3, 7):
    row = ws[i]
    student_num = row[0].value
    student_name = row[1].value or ""
    roll_no = row[2].value or ""
    marks_secured = row[23].value or "N/A"
    max_marks = row[24].value or "N/A"
    
    print(f"Row {i-2}:")
    print(f"  #: {student_num}")
    print(f"  Name: {student_name}")
    print(f"  Roll: {roll_no}")
    print(f"  Marks Secured: {marks_secured}")
    print(f"  Max Marks: {max_marks}")
    print()

wb.close()
