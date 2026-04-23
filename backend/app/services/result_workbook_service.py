"""Service for managing result.xlsx file operations."""
from pathlib import Path
import re
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment


# Column mapping for questions: Q1a -> D, Q1b -> E, Q2a -> F, etc.
QUESTION_COLUMN_MAP = {
    "Q1a": 4,   # D
    "Q1b": 5,   # E
    "Q2a": 6,   # F
    "Q2b": 7,   # G
    "Q3a": 8,   # H
    "Q3b": 9,   # I
    "Q4a": 10,  # J
    "Q4b": 11,  # K
    "Q5a": 12,  # L
    "Q5b": 13,  # M
    "Q6a": 14,  # N
    "Q6b": 15,  # O
    "Q7a": 16,  # P
    "Q7b": 17,  # Q
    "Q8a": 18,  # R
    "Q8b": 19,  # S
    "Q9a": 20,  # T
    "Q9b": 21,  # U
    "Q10a": 22, # V
    "Q10b": 23, # W
}

MARKS_SECURED_COLUMN = 24  # X
MAX_MARKS_COLUMN = 25      # Y


def _parse_numeric_mark(value: str | int | float | None) -> float | None:
    if value in (None, ""):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return None

    try:
        parsed = float(match.group(0))
    except ValueError:
        return None

    if parsed < 0 or parsed > 10:
        return None

    return parsed


def get_or_create_result_workbook(output_path: str | Path) -> Path:
    """Create result.xlsx if it doesn't exist with the proper structure."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    if output_path.exists():
        return output_path
    
    from openpyxl import Workbook
    
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Results"
    
    # Row 1: Question headers
    row1 = ["#", "Student Name", "Roll No"]
    for q_num in range(1, 11):
        row1.append(f"Q{q_num}")
        row1.append(None)
    row1.extend(["Marks Secured", "Max Marks"])
    
    for col_idx, value in enumerate(row1, start=1):
        ws.cell(row=1, column=col_idx, value=value)
    
    # Row 2: Subquestion headers (a, b)
    row2 = [None, None, None]
    for _ in range(10):
        row2.extend(["a", "b"])
    row2.extend([None, None])
    
    for col_idx, value in enumerate(row2, start=1):
        ws.cell(row=2, column=col_idx, value=value)
    
    # Add one empty row for data entry
    ws.cell(row=3, column=1, value=1)
    
    workbook.save(output_path)
    return output_path


def update_result_workbook(
    result_file: str | Path,
    row_num: int,
    student_name: str,
    roll_no: str,
    question_marks: dict[str, str | int | float],
    total_marks: Optional[str | int | float] = None,
) -> Path:
    """Update a row in result.xlsx with extracted question marks."""
    result_file = Path(result_file)
    
    if not result_file.exists():
        get_or_create_result_workbook(result_file)
    
    workbook = load_workbook(result_file)
    ws = workbook.active
    
    # Calculate actual row in sheet (accounting for header rows)
    sheet_row = row_num + 2  # +2 for header rows
    
    # Ensure the row exists
    if sheet_row > ws.max_row:
        for r in range(ws.max_row + 1, sheet_row + 1):
            ws.cell(row=r, column=1, value=r - 2)
    
    # Set the data
    ws.cell(row=sheet_row, column=1, value=row_num)
    ws.cell(row=sheet_row, column=2, value=student_name)
    ws.cell(row=sheet_row, column=3, value=roll_no)
    
    # Populate question marks and calculate total
    total_marks_secured = 0
    for qkey, mark in question_marks.items():
        if qkey in QUESTION_COLUMN_MAP:
            col_idx = QUESTION_COLUMN_MAP[qkey]
            mark_value = str(mark).strip() if mark not in (None, "") else ""
            ws.cell(row=sheet_row, column=col_idx, value=mark_value)
            
            # Sum up marks for marks secured
            parsed = _parse_numeric_mark(mark_value)
            if parsed is not None:
                total_marks_secured += parsed
    
    # Set Marks Secured (sum of all question marks)
    ws.cell(row=sheet_row, column=MARKS_SECURED_COLUMN, value=int(total_marks_secured))
    
    # Set Max Marks to fixed value of 50
    ws.cell(row=sheet_row, column=MAX_MARKS_COLUMN, value=50)
    
    workbook.save(result_file)
    return result_file


def read_result_workbook_row(result_file: str | Path, row_num: int) -> dict[str, Any]:
    """Read a specific row from result.xlsx."""
    result_file = Path(result_file)
    
    if not result_file.exists():
        return {}
    
    workbook = load_workbook(result_file, read_only=True, data_only=True)
    ws = workbook.active
    
    sheet_row = row_num + 2
    if sheet_row > ws.max_row:
        return {}
    
    row_data = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=sheet_row, column=col_idx).value
        row_data[col_idx] = cell_value
    
    return row_data
