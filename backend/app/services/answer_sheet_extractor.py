from __future__ import annotations

import base64
import json
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import cv2
import pytesseract
import google.generativeai as genai
from openpyxl import load_workbook
from PIL import Image

from app.config import TESSERACT_PATH
from app.services.excel_service import write_extraction_workbook


QUESTION_KEYS = [f"Q{q_num}{part}" for q_num in range(1, 11) for part in ("a", "b")]


QUESTIONS_PROMPT_TEMPLATE = """Read the marks table and extract marks for each key from Q1a to Q10b.

Important rules:
1. Match by the printed question labels (Q1..Q10 and a/b), not by guessing.
2. Return only numeric marks (for example 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10) or empty string.
3. If unsure for a key, return empty string for that key.

Return JSON with exactly these keys only:
{"Q1a":"","Q1b":"","Q2a":"","Q2b":"","Q3a":"","Q3b":"","Q4a":"","Q4b":"","Q5a":"","Q5b":"","Q6a":"","Q6b":"","Q7a":"","Q7b":"","Q8a":"","Q8b":"","Q9a":"","Q9b":"","Q10a":"","Q10b":""}
"""

ORDERED_MARKS_PROMPT_TEMPLATE = """Read only the handwritten marks entry cells from the marks table.

Return JSON with a single key "ordered_marks" containing exactly 20 values in this strict order:
[Q1a, Q1b, Q2a, Q2b, Q3a, Q3b, Q4a, Q4b, Q5a, Q5b, Q6a, Q6b, Q7a, Q7b, Q8a, Q8b, Q9a, Q9b, Q10a, Q10b]

Use only numeric strings or empty string for each value. Do not add any other keys.
"""

SINGLE_CELL_PROMPT_TEMPLATE = """This image is one marks cell for a single question part.
Extract only the mark written in this cell.
Return JSON only as: {"mark":""}
Use numeric value string if visible, else empty string.
"""

BACKEND_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = BACKEND_ROOT / "data"
REFERENCE_WORKBOOK = DATA_DIR / "data.xlsx"
DEFAULT_OUTPUT_DIR = DATA_DIR / "outputs"
DEFAULT_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")

pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH


@dataclass
class ExtractionResult:
    roll_no: str
    student_name: str
    marks_rows: list[dict[str, Any]]
    output_path: Path
    source_image: Path
    reference_workbook: Path
    extraction_mode: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "roll_no": self.roll_no,
            "student_name": self.student_name,
            "marks_rows": self.marks_rows,
            "output_path": str(self.output_path),
            "source_image": str(self.source_image),
            "reference_workbook": str(self.reference_workbook),
            "extraction_mode": self.extraction_mode,
        }


def _load_image(image_path: str | Path) -> Any:
    image = cv2.imread(str(image_path))
    if image is None:
        raise ValueError(f"Unable to read image: {image_path}")
    return image


def split_answer_sheet_regions(image: Any) -> dict[str, Any]:
    height, width = image.shape[:2]
    roll_bottom = max(int(height * 0.36), 1)
    overlap = max(int(height * 0.04), 32)

    return {
        "roll_region": image[:roll_bottom, :width],
        "marks_region": image[max(0, roll_bottom - overlap):height, :width],
    }


def _encode_image(image: Any) -> str:
    success, buffer = cv2.imencode(".jpg", image)
    if not success:
        raise ValueError("Failed to encode image for Gemini request")
    return base64.b64encode(buffer).decode("ascii")


def _build_model(api_key: str | None = None, model_name: str = DEFAULT_MODEL):
    key = api_key or "YOUR_API_KEY_HERE"

    genai.configure(api_key=key)
    return genai.GenerativeModel(model_name)


def _strip_json(text: str) -> str:
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?", "", cleaned, flags=re.IGNORECASE).strip()
    cleaned = re.sub(r"```$", "", cleaned).strip()
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start != -1 and end != -1 and end > start:
        cleaned = cleaned[start:end + 1]
    return cleaned


def _parse_json_payload(text: str) -> dict[str, Any]:
    cleaned = _strip_json(text)
    payload = json.loads(cleaned)
    if not isinstance(payload, dict):
        raise ValueError("Gemini response did not return a JSON object")
    return payload


def _call_gemini_json(model: Any, image: Any, prompt: str) -> tuple[dict[str, Any], str]:
    pil_image = Image.fromarray(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
    response = model.generate_content(
        [prompt, pil_image],
        generation_config={
            "temperature": 0,
            "response_mime_type": "application/json",
        },
    )

    text = getattr(response, "text", "") or ""
    if not text:
        raise ValueError("Gemini response did not contain text output")
    return _parse_json_payload(text), text


def _ocr_text(image: Any) -> str:
    pil_image = Image.fromarray(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
    return pytesseract.image_to_string(pil_image, config="--oem 3 --psm 6").strip()


def _extract_line_positions(line_image: Any, axis: str) -> list[int]:
    if axis == "x":
        projection = line_image.sum(axis=0)
        threshold = max(1000, int(line_image.shape[0] * 255 * 0.20))
    else:
        projection = line_image.sum(axis=1)
        threshold = max(1000, int(line_image.shape[1] * 255 * 0.20))

    raw_positions = [idx for idx, value in enumerate(projection) if value >= threshold]
    if not raw_positions:
        return []

    clustered: list[list[int]] = [[raw_positions[0]]]
    for pos in raw_positions[1:]:
        if pos - clustered[-1][-1] <= 2:
            clustered[-1].append(pos)
        else:
            clustered.append([pos])

    return [int(sum(group) / len(group)) for group in clustered]


def _extract_table_roi(marks_region: Any) -> Any:
    gray = cv2.cvtColor(marks_region, cv2.COLOR_BGR2GRAY)
    bin_img = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        31,
        15,
    )

    h, w = gray.shape[:2]
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(32, w // 3), 1))
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(24, h // 10)))
    h_lines = cv2.morphologyEx(bin_img, cv2.MORPH_OPEN, h_kernel)
    v_lines = cv2.morphologyEx(bin_img, cv2.MORPH_OPEN, v_kernel)
    grid = cv2.add(h_lines, v_lines)

    contours, _ = cv2.findContours(grid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    candidates: list[tuple[int, int, int, int]] = []

    for cnt in contours:
        x, y, cw, ch = cv2.boundingRect(cnt)
        if cw >= int(0.65 * w) and ch >= int(0.18 * h):
            candidates.append((x, y, cw, ch))

    if not candidates:
        return marks_region

    # Prefer the widest candidate closest to the top of marks_region.
    candidates.sort(key=lambda box: (-box[2], box[1]))
    x, y, cw, ch = candidates[0]
    return marks_region[y:y + ch, x:x + cw]


def _ocr_cell_mark(cell_image: Any) -> str:
    if cell_image is None or cell_image.size == 0:
        return ""

    gray = cv2.cvtColor(cell_image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (3, 3), 0)
    _, thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    inv = 255 - thresh
    upscaled = cv2.resize(inv, None, fx=2.0, fy=2.0, interpolation=cv2.INTER_CUBIC)

    text = pytesseract.image_to_string(
        upscaled,
        config="--oem 3 --psm 10 -c tessedit_char_whitelist=0123456789",
    )
    return _normalize_mark_value(text)


def _fallback_question_marks(marks_region: Any) -> dict[str, str]:
    table = _extract_table_roi(marks_region)
    gray = cv2.cvtColor(table, cv2.COLOR_BGR2GRAY)
    bin_img = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        31,
        15,
    )

    h, w = gray.shape[:2]
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(24, w // 5), 1))
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(16, h // 8)))
    h_lines = cv2.morphologyEx(bin_img, cv2.MORPH_OPEN, h_kernel)
    v_lines = cv2.morphologyEx(bin_img, cv2.MORPH_OPEN, v_kernel)

    y_lines = _extract_line_positions(h_lines, axis="y")
    x_lines = _extract_line_positions(v_lines, axis="x")

    if len(y_lines) >= 6:
        a_top, a_bottom = y_lines[3], y_lines[4]
        b_top, b_bottom = y_lines[4], y_lines[5]
    else:
        a_top, a_bottom = int(0.50 * h), int(0.64 * h)
        b_top, b_bottom = int(0.64 * h), int(0.78 * h)

    if len(x_lines) >= 12:
        q_boundaries = x_lines[-12:]
    else:
        left = int(0.16 * w)
        right = int(0.98 * w)
        step = (right - left) / 11.0
        q_boundaries = [int(left + idx * step) for idx in range(12)]

    question_marks = {qkey: "" for qkey in QUESTION_KEYS}
    y_pad = 2
    x_pad = 2

    for q_num in range(1, 11):
        col_idx = q_num - 1
        x0 = max(0, min(q_boundaries[col_idx] + x_pad, w - 1))
        x1 = max(x0 + 1, min(q_boundaries[col_idx + 1] - x_pad, w))

        a0 = max(0, min(a_top + y_pad, h - 1))
        a1 = max(a0 + 1, min(a_bottom - y_pad, h))
        b0 = max(0, min(b_top + y_pad, h - 1))
        b1 = max(b0 + 1, min(b_bottom - y_pad, h))

        cell_a = table[a0:a1, x0:x1]
        cell_b = table[b0:b1, x0:x1]

        question_marks[f"Q{q_num}a"] = _ocr_cell_mark(cell_a)
        question_marks[f"Q{q_num}b"] = _ocr_cell_mark(cell_b)

    return question_marks


def _fallback_roll_no(image: Any) -> str:
    text = _ocr_text(image)
    match = re.search(
        r"(?:roll\s*no|roll\s*number|register\s*no|reg\s*no)\s*[:\-]?\s*([A-Za-z0-9/-]+)",
        text,
        flags=re.IGNORECASE,
    )
    if match:
        return _normalize_roll_no(match.group(1))

    token_match = re.search(r"[A-Za-z0-9/-]{3,}", text)
    return _normalize_roll_no(token_match.group(0)) if token_match else ""


def _fallback_marks_rows(image: Any) -> list[dict[str, Any]]:
    text = _ocr_text(image)
    rows: list[dict[str, Any]] = []
    for line_no, line in enumerate((line.strip() for line in text.splitlines() if line.strip()), start=1):
        rows.append({"row_no": line_no, "raw_text": line})
    return rows


def _normalize_roll_no(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("Roll No", "").replace("Roll Number", "")
    text = text.replace(":", " ")
    text = re.sub(r"\s+", "", text)
    return text


def _normalize_name(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_mark_value(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""

    # Keep only the first numeric token to avoid values like "2 marks" or "2/3".
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return ""

    candidate = match.group(0)
    try:
        numeric = float(candidate)
    except ValueError:
        return ""

    # Per-question marks are expected to be in a small range.
    if numeric < 0 or numeric > 10:
        return ""

    if numeric.is_integer():
        return str(int(numeric))
    return str(numeric)


def _to_question_mark_dict(payload: dict[str, Any]) -> dict[str, str]:
    question_marks: dict[str, str] = {}
    for qkey in QUESTION_KEYS:
        question_marks[qkey] = _normalize_mark_value(payload.get(qkey, ""))
    return question_marks


def _extract_ordered_marks(
    gemini_model: Any,
    marks_region: Any,
) -> dict[str, str]:
    payload, _ = _call_gemini_json(gemini_model, marks_region, ORDERED_MARKS_PROMPT_TEMPLATE)
    ordered = payload.get("ordered_marks")
    if not isinstance(ordered, list):
        return {qkey: "" for qkey in QUESTION_KEYS}

    normalized = [_normalize_mark_value(v) for v in ordered[: len(QUESTION_KEYS)]]
    if len(normalized) < len(QUESTION_KEYS):
        normalized.extend([""] * (len(QUESTION_KEYS) - len(normalized)))

    return {qkey: normalized[idx] for idx, qkey in enumerate(QUESTION_KEYS)}


def _extract_single_cell_mark(gemini_model: Any, cell_image: Any) -> str:
    try:
        payload, _ = _call_gemini_json(gemini_model, cell_image, SINGLE_CELL_PROMPT_TEMPLATE)
        return _normalize_mark_value(payload.get("mark", ""))
    except Exception:
        return ""


def _get_estimated_mark_cells(marks_region: Any) -> dict[str, Any]:
    h, w = marks_region.shape[:2]
    # Focus on the lower band where handwritten marks usually appear.
    y_start = int(h * 0.50)
    y_end = int(h * 0.98)
    y_start = max(0, min(y_start, h - 1))
    y_end = max(y_start + 1, min(y_end, h))

    band = marks_region[y_start:y_end, :]
    bw = band.shape[1]
    cells: dict[str, Any] = {}

    for idx, qkey in enumerate(QUESTION_KEYS):
        x0 = int((idx / len(QUESTION_KEYS)) * bw)
        x1 = int(((idx + 1) / len(QUESTION_KEYS)) * bw)
        x0 = max(0, min(x0, bw - 1))
        x1 = max(x0 + 1, min(x1, bw))
        cells[qkey] = band[:, x0:x1]

    return cells


def load_roll_name_lookup(workbook_path: str | Path = REFERENCE_WORKBOOK) -> dict[str, str]:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Reference workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    roll_index = None
    name_index = None
    for index, header in enumerate(headers):
        if header in {"roll no", "roll number", "register no", "register number", "reg no", "reg number"}:
            roll_index = index
        if header in {"name", "student name"}:
            name_index = index

    if roll_index is None or name_index is None:
        raise ValueError("Unable to locate Roll No and Name columns in the reference workbook")

    lookup: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        roll_value = _normalize_roll_no(row[roll_index] if roll_index < len(row) else None)
        name_value = _normalize_name(row[name_index] if name_index < len(row) else None)
        if roll_value:
            lookup[roll_value] = name_value

    return lookup


def _normalize_marks_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    rows = payload.get("marks_rows") or payload.get("rows") or payload.get("table_rows") or []
    normalized_rows: list[dict[str, Any]] = []

    for index, row in enumerate(rows, start=1):
        if isinstance(row, dict):
            normalized_row = {str(key): value for key, value in row.items() if value not in (None, "")}
        else:
            normalized_row = {"value": row}
        normalized_row["row_no"] = index
        normalized_rows.append(normalized_row)

    return normalized_rows


def extract_answer_sheet(
    image_path: str | Path,
    output_path: str | Path | None = None,
    reference_workbook: str | Path = REFERENCE_WORKBOOK,
    api_key: str | None = None,
    model_name: str = DEFAULT_MODEL,
) -> ExtractionResult:
    image_path = Path(image_path)
    image = _load_image(image_path)
    regions = split_answer_sheet_regions(image)

    gemini_model = _build_model(api_key, model_name)
    extraction_mode = "gemini"

    roll_prompt = (
        "Extract only the roll number from this answer sheet region. "
        'Return JSON only in the form {"roll_no":"..."}. '
        "If the roll number is not present, use an empty string."
    )
    marks_prompt = (
        "Extract the marks table from this answer sheet region. "
        'Return JSON only in the form {"marks_rows":[...]}. '
        "Preserve the row order and keep values as strings. "
        "Use concise keys such as subject, marks, max_marks, remarks when they are visible."
    )

    try:
        roll_payload, _ = _call_gemini_json(gemini_model, regions["roll_region"], roll_prompt)
        marks_payload, _ = _call_gemini_json(gemini_model, regions["marks_region"], marks_prompt)
        roll_no = _normalize_roll_no(roll_payload.get("roll_no", ""))
        marks_rows = _normalize_marks_rows(marks_payload)
    except Exception:
        extraction_mode = "ocr_fallback"
        roll_no = _fallback_roll_no(regions["roll_region"])
        marks_rows = _fallback_marks_rows(regions["marks_region"])

    lookup = load_roll_name_lookup(reference_workbook)
    student_name = lookup.get(roll_no, "")

    if output_path is None:
        output_path = DEFAULT_OUTPUT_DIR / f"{image_path.stem}_extracted.xlsx"

    summary = {
        "source_image": image_path.name,
        "roll_no": roll_no,
        "student_name": student_name,
        "reference_workbook": str(Path(reference_workbook)),
        "api_provider": extraction_mode,
        "model_name": model_name,
        "extraction_mode": extraction_mode,
        "marks_rows_count": len(marks_rows),
    }

    saved_path = write_extraction_workbook(output_path, summary, marks_rows)

    return ExtractionResult(
        roll_no=roll_no,
        student_name=student_name,
        marks_rows=marks_rows,
        output_path=saved_path,
        source_image=image_path,
        reference_workbook=Path(reference_workbook),
        extraction_mode=extraction_mode,
    )


def _extract_question_marks(
    gemini_model: Any,
    marks_region: Any,
) -> dict[str, str]:
    """Extract marks with consistency checks to reduce value and mapping errors."""
    try:
        keyed_payload, _ = _call_gemini_json(gemini_model, marks_region, QUESTIONS_PROMPT_TEMPLATE)
        keyed_marks = _to_question_mark_dict(keyed_payload)
    except Exception:
        keyed_marks = {qkey: "" for qkey in QUESTION_KEYS}

    try:
        ordered_marks = _extract_ordered_marks(gemini_model, marks_region)
    except Exception:
        ordered_marks = {qkey: "" for qkey in QUESTION_KEYS}

    final_marks: dict[str, str] = {}
    conflicting_keys: list[str] = []

    for qkey in QUESTION_KEYS:
        keyed = keyed_marks.get(qkey, "")
        ordered = ordered_marks.get(qkey, "")

        if keyed and ordered:
            if keyed == ordered:
                final_marks[qkey] = keyed
            else:
                # Defer disagreement to per-cell read.
                final_marks[qkey] = keyed
                conflicting_keys.append(qkey)
        elif keyed:
            final_marks[qkey] = keyed
        else:
            final_marks[qkey] = ordered

    if conflicting_keys:
        cell_map = _get_estimated_mark_cells(marks_region)
        for qkey in conflicting_keys:
            cell_img = cell_map.get(qkey)
            if cell_img is None or cell_img.size == 0:
                continue
            single_cell_mark = _extract_single_cell_mark(gemini_model, cell_img)
            if single_cell_mark:
                final_marks[qkey] = single_cell_mark
            elif not final_marks.get(qkey):
                final_marks[qkey] = ordered_marks.get(qkey, "")

    for qkey in QUESTION_KEYS:
        final_marks[qkey] = _normalize_mark_value(final_marks.get(qkey, ""))

    return final_marks


def extract_to_result_workbook(
    image_path: str | Path,
    result_workbook_path: str | Path = None,
    reference_workbook: str | Path = REFERENCE_WORKBOOK,
    api_key: str | None = None,
    model_name: str = DEFAULT_MODEL,
    row_num: int = 1,
) -> dict[str, Any]:
    """Extract marks from answer sheet image and write to result.xlsx."""
    from app.services.result_workbook_service import (
        get_or_create_result_workbook,
        update_result_workbook,
    )

    image_path = Path(image_path)
    
    if result_workbook_path is None:
        result_workbook_path = DEFAULT_OUTPUT_DIR / "result.xlsx"
    
    result_workbook_path = Path(result_workbook_path)
    result_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Get or create result.xlsx
    get_or_create_result_workbook(result_workbook_path)
    
    # Load image
    image = _load_image(image_path)
    
    # Setup Gemini model
    gemini_model = _build_model(api_key, model_name)
    extraction_mode = "gemini"
    
    # Define unified prompt exactly as requested
    combined_prompt = (
        "Extract the roll number from the image.\n"
        "And extract the marks from the marks table with its corresponding qn number and subqn no.\n"
        "Context: the roll no extracted is then used to match the name from the data.xlsx and the generate the final output using all these data.\n"
        "Return the output strictly in JSON format with exactly this structure:\n"
        '{"roll_no": "", "marks": {"Q1a":"","Q1b":"","Q2a":"","Q2b":"","Q3a":"","Q3b":"","Q4a":"","Q4b":"","Q5a":"","Q5b":"","Q6a":"","Q6b":"","Q7a":"","Q7b":"","Q8a":"","Q8b":"","Q9a":"","Q9b":"","Q10a":"","Q10b":""}}\n'
        "Use empty strings for empty or unreadable marks."
    )
    
    try:
        # Pass the entire image to Gemini
        payload, _ = _call_gemini_json(gemini_model, image, combined_prompt)
        roll_no = _normalize_roll_no(payload.get("roll_no", ""))
        
        # Extract question-level marks from unified JSON response
        question_marks = {}
        raw_marks = payload.get("marks", {})
        if isinstance(raw_marks, dict):
            for qkey in QUESTION_KEYS:
                question_marks[qkey] = _normalize_mark_value(raw_marks.get(qkey, ""))
        else:
            for qkey in QUESTION_KEYS:
                question_marks[qkey] = ""

        # Pre-process regions for fallback
        regions = split_answer_sheet_regions(image)
        ocr_marks = _fallback_question_marks(regions["marks_region"])

        # Reconcile: prefer OCR only when Gemini left blank for that key.
        for qkey in QUESTION_KEYS:
            if not question_marks.get(qkey) and ocr_marks.get(qkey):
                question_marks[qkey] = ocr_marks[qkey]
        if not roll_no:
            roll_no = _fallback_roll_no(regions["roll_region"])
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise RuntimeError(f"Gemini API Extraction Failed: {str(e)}")
    
    # Look up student name
    lookup = load_roll_name_lookup(reference_workbook)
    student_name = lookup.get(roll_no, "")
    
    # Update result workbook (marks will be calculated internally)
    update_result_workbook(
        result_workbook_path,
        row_num=row_num,
        student_name=student_name,
        roll_no=roll_no,
        question_marks=question_marks,
    )
    
    # Calculate total for return value
    total_marks_for_return = 0
    for mark_val in question_marks.values():
        if mark_val and str(mark_val)[0].isdigit():
            try:
                total_marks_for_return += float(mark_val)
            except (ValueError, TypeError):
                pass
    
    return {
        "status": "success",
        "extraction_mode": extraction_mode,
        "roll_no": roll_no,
        "student_name": student_name,
        "question_marks": question_marks,
        "total_marks": total_marks_for_return,
        "result_file": str(result_workbook_path),
        "source_image": str(image_path),
    }