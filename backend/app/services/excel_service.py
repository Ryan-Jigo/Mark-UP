from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook


def _stringify(value: Any) -> str:
	if value is None:
		return ""
	return str(value)


def write_extraction_workbook(
	output_path: str | Path,
	summary: dict[str, Any],
	marks_rows: Iterable[dict[str, Any]],
) -> Path:
	output_path = Path(output_path)
	output_path.parent.mkdir(parents=True, exist_ok=True)

	workbook = Workbook()
	summary_sheet = workbook.active
	summary_sheet.title = "Summary"

	for row_index, (key, value) in enumerate(summary.items(), start=1):
		summary_sheet.cell(row=row_index, column=1, value=key)
		summary_sheet.cell(row=row_index, column=2, value=_stringify(value))

	marks_sheet = workbook.create_sheet("Marks")
	normalized_rows = [
		{key: _stringify(value) for key, value in row.items()}
		for row in marks_rows
	]

	headers = ["row_no"]
	for row in normalized_rows:
		for key in row.keys():
			if key not in headers:
				headers.append(key)

	for column_index, header in enumerate(headers, start=1):
		marks_sheet.cell(row=1, column=column_index, value=header)

	for row_index, row in enumerate(normalized_rows, start=2):
		marks_sheet.cell(row=row_index, column=1, value=row_index - 1)
		for column_index, header in enumerate(headers[1:], start=2):
			marks_sheet.cell(row=row_index, column=column_index, value=row.get(header, ""))

	workbook.save(output_path)
	return output_path
